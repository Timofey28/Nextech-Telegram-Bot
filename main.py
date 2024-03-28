import os
from datetime import time, datetime
import pytz
from typing import Optional, Tuple
import logging
import pandas as pd
import openpyxl as xl
from openpyxl.styles import Alignment, GradientFill
from telegram import (
    Chat,
    ChatMember,
    ChatMemberUpdated,
    Update,
    KeyboardButton,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    BotCommandScopeAllPrivateChats,
    BotCommandScopeChat,
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
    ChatMemberHandler,
    CallbackQueryHandler,
    CallbackContext,
    Defaults,
)
from data import TOKEN, db_dbname, db_host, db_user, db_password
from database import Database
from debt_display_handler import DebtDisplayHandler
from debt_display_strategies.aggregation_by_dates import AggregationByDates
from debt_display_strategies.aggregation_by_people import AggregationByPeople
from clients import Clients
from contracts import Contracts
from payments import Payments
from global_vars import GlobalVars


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_type = update.message.chat.type
    chat_id = update.effective_chat.id
    msg = update.message.text

    if chat_type == 'private':
        try:
            await context.bot.send_message(gv.ACCOUNTANT_GROUP_ID, f'Мне тут {update.effective_user.full_name} что-то написал, не умею читать ( ⬇⬇')
            await context.bot.forward_message(gv.ACCOUNTANT_GROUP_ID, update.effective_user.id, update.message.id)
        except:
            logging.error('не получилось переслать сообщение от клиента, так как ADMIN_GROUP_ID == 0; возможно, отсутствует файл "admin_group_id.txt"')
    elif chat_type == 'group':
        if chat_id == gv.ACCOUNTANT_GROUP_ID:
            if gv.ACCOUNTANT_GROUP_STATUS == gv.Status_ENTERING_BUSINESS_DIRECTION:
                await clients.enter_business_direction(update, context)

            elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_ENTERING_CLIENTS_DATA:
                await clients.enter_client_data(update, context)

            elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_EDITING_CLIENT:
                GROUP_ID = gv.ACCOUNTANT_GROUP_ID
                if msg == "отмена":
                    gv.set_group_status('accountant', gv.Status_IN_SECTION_CLIENTS)
                    await context.bot.send_message(GROUP_ID, 'Хорошо 😉', reply_markup=ReplyKeyboardRemove())
                else:
                    result = clients.check_entered_client_data_and_update(msg, gv.CALLBACK_DATA['client_id'])
                    if type(result) is str:
                        await context.bot.send_message(GROUP_ID, result)
                    else:
                        gv.set_group_status('accountant', gv.Status_IN_SECTION_CLIENTS)
                        if result:
                            await context.bot.send_message(GROUP_ID, 'Перезаписал данные клиента ✅', reply_markup=ReplyKeyboardRemove())
                        else:
                            await context.bot.send_message(GROUP_ID, 'Не удалось перезаписать данные клиента ⚠', reply_markup=ReplyKeyboardRemove())

            elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_SENDING_NEW_CONTRACT_TEMPLATE:
                gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
                await context.bot.send_message(gv.ACCOUNTANT_GROUP_ID, 'Ок', reply_markup=ReplyKeyboardRemove())

            elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_ENTERING_CONTRACT_INFO_STAGE_1:
                await contracts.enter_contract_data_stage_1(update, context)

            elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_ENTERING_CONTRACT_INFO_STAGE_2:
                GROUP_ID = gv.ACCOUNTANT_GROUP_ID
                if msg == "отмена":
                    gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
                    await context.bot.send_message(GROUP_ID, '🆗', reply_markup=ReplyKeyboardRemove())

            elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_ENTERING_CONTRACT_INFO_STAGE_3:
                await contracts.enter_contract_data_stage_3(update, context)

            elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_SENDING_UPDATED_CONTRACT:
                GROUP_ID = gv.ACCOUNTANT_GROUP_ID
                if msg == 'отмена':
                    gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
                    await context.bot.send_message(GROUP_ID, '🆗', reply_markup=ReplyKeyboardRemove())

            elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_ENTERING_UPDATED_CONTRACT_PAYMENT_SCHEDULE:
                await contracts.enter_updated_contract_payment_schedule(update, context)

            elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_ENTERING_PAID_PAYMENT_INFO:
                await payments.enter_paid_payment_info(update, context)

    else:
        logging.warning(f'сообщение в чате типа: {chat_type}')

    print(f'User ({update.message.chat.id}) in {chat_type}: "{msg}"')


async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query

    if gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_BUSINESS_DIRECTION_TO_DELETE_ONE:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            await context.bot.send_message(GROUP_ID, "Как скажешь")
        else:
            chosen_direction = gv.CALLBACK_DATA[query.data]
            result = db.delete_business_direction(chosen_direction)
            if result is True:
                await context.bot.send_message(GROUP_ID, f'Удалил направление "{chosen_direction}" 👍')
            else:
                await context.bot.send_message(GROUP_ID, f'Не могу удалить направление бизнеса "{chosen_direction}", так как в нем находится как минимум один человек 😐')
        gv.set_group_status('accountant', gv.Status_IN_SECTION_CLIENTS)

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_BUSINESS_DIRECTION_TO_ADD_NEW_CLIENT:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CLIENTS)
            await context.bot.send_message(GROUP_ID, "Да легко")
        else:
            gv.set_group_status('accountant', gv.Status_ENTERING_CLIENTS_DATA)
            gv.CALLBACK_DATA['chosen_direction'] = gv.CALLBACK_DATA[query.data]
            reply_msg = f'Было выбрано направление "{gv.CALLBACK_DATA[query.data]}". Введи данные нового клиента в следующем формате:\n\n{clients.EXPECTED_CLIENTS_DATA_FORMAT_TO_ADD}\n\nили нажми "отмена"'
            await context.bot.send_message(GROUP_ID, reply_msg, reply_markup=ReplyKeyboardMarkup(gv.CANCEL_BUTTON))

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_BUSINESS_DIRECTION_TO_SEND_FILE:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            os.remove(gv.CURRENT_FILE)
            await context.bot.send_message(GROUP_ID, "🆗")
            gv.set_group_status('accountant', gv.Status_MAIN)
        else:
            gv.set_group_status('accountant', gv.Status_PRESSING_CLIENT_TO_SEND_FILE)
            buttons = []
            clients_ = db.get_clients_from_business_direction(gv.CALLBACK_DATA[query.data])
            gv.CALLBACK_DATA.clear()
            for i in range(len(clients_)):
                name = clients_[i]['name']
                surname = clients_[i]['surname']
                tg_id = clients_[i]['tg_id']
                phone_number = clients_[i]['phone_number']
                buttons.append([InlineKeyboardButton(f'{name} {surname}, {prettify_phone_number(phone_number)}', callback_data=f'{tg_id}')])
                gv.CALLBACK_DATA[f'{tg_id}'] = f'{name} {surname}'
            buttons.append([InlineKeyboardButton("никому, забей", callback_data="отмена")])
            await context.bot.send_message(GROUP_ID, 'Кому из них отправить?', reply_markup=InlineKeyboardMarkup(buttons))

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_CLIENT_TO_SEND_FILE:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            await context.bot.send_message(GROUP_ID, "🆒")
        else:
            assert gv.CURRENT_FILE != ''
            try:
                sent_message = await context.bot.send_document(query.data, gv.CURRENT_FILE)
                db.write_sent_message_info(int(query.data), sent_message.message_id)
                await context.bot.send_message(GROUP_ID, f'Скинул файл клиенту по имени {gv.CALLBACK_DATA[query.data]} ✅')
            except:
                await context.bot.send_message(GROUP_ID, f'❌ Не удалось скинуть файл человеку по имени {gv.CALLBACK_DATA[query.data]}')
        os.remove(gv.CURRENT_FILE)
        gv.set_group_status('accountant', gv.Status_MAIN)

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_BUSINESS_DIRECTION_TO_DELETE_CLIENT:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            await context.bot.send_message(GROUP_ID, "Ладно, как скажешь 😉")
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CLIENTS)
        else:
            gv.set_group_status('accountant', gv.Status_PRESSING_CLIENT_TO_DELETE_ONE)
            buttons = []
            clients_ = db.get_clients_from_business_direction(gv.CALLBACK_DATA[query.data])
            gv.CALLBACK_DATA.clear()
            for i in range(len(clients_)):
                name = clients_[i]['name']
                surname = clients_[i]['surname']
                tg_id = clients_[i]['tg_id']
                phone_number = clients_[i]['phone_number']
                buttons.append([InlineKeyboardButton(f'{name} {surname}, {prettify_phone_number(phone_number)}', callback_data=f'{tg_id}')])
                gv.CALLBACK_DATA[f'{tg_id}'] = f'{name} {surname}'
            buttons.append([InlineKeyboardButton("никого не надо, забей", callback_data="отмена")])
            await context.bot.send_message(GROUP_ID, 'Кого из них удалить из базы?', reply_markup=InlineKeyboardMarkup(buttons))

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_CLIENT_TO_DELETE_ONE:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            await context.bot.send_message(GROUP_ID, "Хорошо 😉")
        else:
            db.delete_client(int(query.data))
            await context.bot.send_message(GROUP_ID, f'Удалил клиента по имени {gv.CALLBACK_DATA[query.data]} ✅')
        gv.set_group_status('accountant', gv.Status_IN_SECTION_CLIENTS)

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_CLIENT_TO_DELETE_LAST_MESSAGE:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            await context.bot.send_message(GROUP_ID, "Не надо так не надо 😉")
        else:
            client_full_name = gv.CALLBACK_DATA[query.data]
            chat_id, message_id = list(map(int, query.data.split(',')))
            try:
                await context.bot.delete_message(chat_id, message_id)
                db.delete_sent_message_info(chat_id, message_id)
                await context.bot.send_message(GROUP_ID, f'Удалил последнее сообщение клиенту по имени {client_full_name} ✅')
            except:
                await context.bot.send_message(GROUP_ID, f'❌ Не удалось удалить последнее сообщение клиенту по имени {client_full_name}')
        gv.set_group_status('accountant', gv.Status_MAIN)

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_BUTTON_WHILE_PAYING_SALARIES:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            await context.bot.send_message(GROUP_ID, 'Хорошо 🗿')
            gv.set_group_status('accountant', gv.Status_MAIN)
        else:
            assert query.data == "пропустить" or query.data == "оплачено"
            if query.data == "пропустить":
                gv.CURRENT_OFFSET += 1
                await context.bot.send_message(GROUP_ID, 'Пока пропускаю, но мы вернемся к этому долгу в конце)')
            else:
                debt_display_handler.mark_shift_as_paid(gv.CURRENT_DEBT)
                await context.bot.send_message(GROUP_ID, 'Кайф 🔥')

            msg, gv.CURRENT_DEBT = debt_display_handler.get_next_payment_message(offset=gv.CURRENT_OFFSET)
            if not msg:
                if gv.CURRENT_OFFSET != 0:
                    gv.CURRENT_OFFSET = 0
                    msg, gv.CURRENT_DEBT = debt_display_handler.get_next_payment_message()
                    if not msg:
                        await context.bot.send_message(GROUP_ID, 'Поздравляю, все задолженности на текущий момент успешно оплачены! ✅🦾')
                        gv.set_group_status('accountant', gv.Status_MAIN)
                        return
                else:
                    await context.bot.send_message(GROUP_ID, 'Поздравляю, все задолженности на текущий момент успешно оплачены! ✅🦾')
                    gv.set_group_status('accountant', gv.Status_MAIN)
                    return
            await context.bot.send_message(GROUP_ID, msg, parse_mode='markdown', reply_markup=InlineKeyboardMarkup(gv.PAY_SALARIES_INLINE_KEYBOARD))

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_BUTTON_TO_DECIDE_WHETHER_TO_PAY_SALARIES:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        assert query.data == "не сейчас" or query.data == "оплатить"
        if query.data == "не сейчас":
            await context.bot.send_message(GROUP_ID, 'Хорошо, но не откладывай это надолго 😉')
            gv.set_group_status('accountant', gv.Status_MAIN)
        else:
            gv.CURRENT_OFFSET = 0
            msg, gv.CURRENT_DEBT = debt_display_handler.get_next_payment_message()
            gv.set_group_status('accountant', gv.Status_PRESSING_BUTTON_WHILE_PAYING_SALARIES)
            await context.bot.send_message(GROUP_ID, 'Начнем с начала ⬇️')
            await context.bot.send_message(GROUP_ID, msg, parse_mode='markdown', reply_markup=InlineKeyboardMarkup(gv.PAY_SALARIES_INLINE_KEYBOARD))

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_IN_SECTION_CLIENTS:
        if query.data == 'add_business_direction':
            await clients.add_business_direction(update, context)
        elif query.data == 'add_client':
            await clients.add_client(update, context)
        elif query.data == 'get_clients_data':
            await clients.get_clients_data(update, context)
        elif query.data == 'edit_client':
            await clients.edit_client_data(update, context)
        elif query.data == 'delete_client':
            await clients.delete_client(update, context)
        elif query.data == 'delete_business_direction':
            await clients.delete_business_direction(update, context)
        elif query.data == 'show_business_directions':
            await clients.show_business_directions(update, context)

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_BUSINESS_DIRECTION_TO_EDIT_CLIENT:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            await context.bot.send_message(GROUP_ID, "Ладно, как скажешь 😉")
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CLIENTS)
        else:
            gv.set_group_status('accountant', gv.Status_PRESSING_CLIENT_TO_EDIT_ONE)
            buttons = []
            clients_ = db.get_clients_from_business_direction(gv.CALLBACK_DATA[query.data])
            gv.CALLBACK_DATA.clear()
            for i in range(len(clients_)):
                idd = clients_[i]['id']
                surname = clients_[i]['surname']
                name = clients_[i]['name']
                phone_number = clients_[i]['phone_number']
                buttons.append([InlineKeyboardButton(f'{name} {surname}, {prettify_phone_number(phone_number)}', callback_data=f'{idd}')])
                gv.CALLBACK_DATA[f'{idd}'] = {
                    'surname': surname,
                    'name': name,
                    'phone_number': phone_number,
                    'patronymic': clients_[i]['patronymic'],
                    'sex': clients_[i]['sex'],
                    'tg_id': clients_[i]['tg_id'],
                    'inn': clients_[i]['inn'],
                    'email': clients_[i]['email'],
                }
            buttons.append([InlineKeyboardButton("никого не надо, забей", callback_data="отмена")])
            await context.bot.send_message(GROUP_ID, 'Кого из них отредактировать?', reply_markup=InlineKeyboardMarkup(buttons))

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_CLIENT_TO_EDIT_ONE:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CLIENTS)
            await context.bot.send_message(GROUP_ID, 'Хорошо 😉')
        else:
            gv.set_group_status('accountant', gv.Status_EDITING_CLIENT)
            print(gv.CALLBACK_DATA)
            client_data = gv.CALLBACK_DATA[query.data]
            msg = 'Текущие данные клиента:\n'
            msg += f"\nФамилия: {client_data['surname']}"
            msg += f"\nИмя: {client_data['name']}"
            msg += f"\nОтчество: {client_data['patronymic']}"
            msg += f"\nПол: {'м' if client_data['sex'] == 'male' else 'ж'}"
            msg += f"\nИНН: {client_data['inn']}"
            msg += f"\nНомер телефона: {prettify_phone_number(client_data['phone_number'])}"
            msg += f"\nПочта: {client_data['email']}"
            await context.bot.send_message(GROUP_ID, msg)
            gv.CALLBACK_DATA['client_id'] = query.data
            msg = f'Введи новые данные в следующем формате:\n\n{clients.EXPECTED_CLIENTS_DATA_FORMAT_TO_EDIT}\n\nили нажми "отмена"'
            await context.bot.send_message(GROUP_ID, msg, reply_markup=ReplyKeyboardMarkup(gv.CANCEL_BUTTON))

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_IN_SECTION_CONTRACTS:
        if query.data == 'send_contract_template':
            await contracts.send_contract_template(update, context)
        elif query.data == 'add_contract':
            await contracts.add_contract(update, context)
        elif query.data == 'list_of_contracts':
            await contracts.list_of_contracts(update, context)
        elif query.data == 'send_contract':
            await contracts.send_contract(update, context)
        elif query.data == 'edit_contract':
            await contracts.edit_contract(update, context)
        elif query.data == 'add_or_change_contract_template':
            await contracts.add_or_change_contract_template(update, context)
        elif query.data == 'delete_contract_template':
            await contracts.delete_contract_template(update, context)

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_CONTRACT_TEMPLATE_TO_SEND_ONE:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
        if query.data == 'отмена':
            await context.bot.send_message(GROUP_ID, 'Ладно 🌚')
        else:
            contract_template_file = gv.CALLBACK_DATA[query.data]
            await context.bot.send_document(GROUP_ID, f'{gv.PATH_CONCTRACT_TEMPLATES}/{contract_template_file}')

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_CONTRACT_TEMPLATE_TO_DELETE_ONE:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
        if query.data == 'отмена':
            await context.bot.send_message(GROUP_ID, 'Принял 🫡')
        else:
            os.remove(f'{gv.PATH_CONCTRACT_TEMPLATES}/{gv.CALLBACK_DATA[query.data]}')
            await context.bot.send_message(GROUP_ID, 'Удалил 👍')

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_BUSINESS_DIRECTION_TO_ADD_NEW_CONTRACT:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
            await context.bot.send_message(GROUP_ID, '🆗')
        else:
            gv.set_group_status('accountant', gv.Status_PRESSING_CLIENT_TO_ADD_NEW_CONTRACT)
            buttons = []
            clients_ = db.get_clients_from_business_direction(gv.CALLBACK_DATA[query.data])
            gv.CALLBACK_DATA.clear()
            for i in range(len(clients_)):
                idd = clients_[i]['id']
                surname = clients_[i]['surname']
                name = clients_[i]['name']
                phone_number = clients_[i]['phone_number']
                buttons.append([InlineKeyboardButton(f'{name} {surname}, {prettify_phone_number(phone_number)}', callback_data=f'{idd}')])
            buttons.append([InlineKeyboardButton("никому не надо, забей", callback_data="отмена")])
            await context.bot.send_message(GROUP_ID, 'Кому добавить договор?', reply_markup=InlineKeyboardMarkup(buttons))

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_CLIENT_TO_ADD_NEW_CONTRACT:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
            await context.bot.send_message(GROUP_ID, '🆗')
        else:
            gv.CALLBACK_DATA['client_id'] = int(query.data)
            gv.set_group_status('accountant', gv.Status_ENTERING_CONTRACT_INFO_STAGE_1)
            msg = 'Данные договора вводятся в 3 этапа. Из любого этапа можно выйти в главное меню, нажав кнопку "отмена". При этом потеряются все ранее введенные данные'
            await context.bot.send_message(GROUP_ID, msg)
            msg = 'Этап 1️⃣. Введи следующие данные:\n\n1) Компания клиента (при отсутствии поставь прочерк "-")'
            msg += '\n2) Дата заключения договора в формате "дд.мм.гг"'
            msg += '\n3) Полная сумма платежа (либо ежемесячная сумма платежа, если договор регулярный)'
            msg += '\n\nСтроки можно оставлять пустыми, тогда они не будут учитываться'
            await context.bot.send_message(GROUP_ID, msg, reply_markup=ReplyKeyboardMarkup(gv.CANCEL_BUTTON))

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_BUSINESS_DIRECTION_TO_EDIT_CONTRACT:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
            await context.bot.send_message(GROUP_ID, '🆗')
        else:
            gv.set_group_status('accountant', gv.Status_PRESSING_CLIENT_TO_EDIT_CONTRACT)
            buttons = []
            clients_ = db.get_clients_having_active_contracts_from_business_direction(gv.CALLBACK_DATA[query.data])
            gv.CALLBACK_DATA.clear()
            for client in clients_:
                idd = client['id']
                surname = client['surname']
                name = client['name']
                phone_number = client['phone_number']
                buttons.append([InlineKeyboardButton(f'{name} {surname}, {prettify_phone_number(phone_number)}', callback_data=f'{idd}')])
            buttons.append([InlineKeyboardButton("ни с кем, забей", callback_data="отмена")])
            await context.bot.send_message(GROUP_ID, 'Договор с кем отредактировать?', reply_markup=InlineKeyboardMarkup(buttons))

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_CLIENT_TO_EDIT_CONTRACT:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
            await context.bot.send_message(GROUP_ID, '🆗')
        else:
            contracts_ = db.get_client_contracts(int(query.data))
            if contracts_ is None:
                logging.error("client can't have no contracts")
                exit(0)
            gv.CALLBACK_DATA.clear()
            if len(contracts_) == 1:
                gv.set_group_status('accountant', gv.Status_CHOOSING_WHAT_IN_CONTRACT_TO_EDIT)
                gv.CALLBACK_DATA['contract_id'] = contracts_[0][0]
                gv.CALLBACK_DATA['contract_file_name'] = contracts_[0][1]
                gv.CALLBACK_DATA['contract_type'] = contracts_[0][2]
                await context.bot.send_message(GROUP_ID, 'Что именно отредактировать?', reply_markup=InlineKeyboardMarkup(contracts.CONTRACT_EDITING_OPTIONS))
            else:
                gv.set_group_status('accountant', gv.Status_CHOOSING_WHICH_CONTRACT_TO_EDIT)
                buttons = []
                for contract in contracts_:
                    idd, name, typee = contract
                    buttons.append([InlineKeyboardButton(name, callback_data=f'{idd}')])
                    gv.CALLBACK_DATA[f'{idd}'] = [name, typee]
                buttons.append([InlineKeyboardButton("никакой, забей", callback_data="отмена")])
                await context.bot.send_message(GROUP_ID, 'Какой из договоров отредактировать?', reply_markup=InlineKeyboardMarkup(buttons))

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_CHOOSING_WHICH_CONTRACT_TO_EDIT:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
            await context.bot.send_message(GROUP_ID, '🆗')
        else:
            gv.set_group_status('accountant', gv.Status_CHOOSING_WHAT_IN_CONTRACT_TO_EDIT)
            gv.CALLBACK_DATA['contract_id'] = int(query.data)
            gv.CALLBACK_DATA['contract_file_name'] = gv.CALLBACK_DATA[query.data][0]
            gv.CALLBACK_DATA['contract_type'] = gv.CALLBACK_DATA[query.data][1]
            await context.bot.send_message(GROUP_ID, 'Что именно отредактировать?', reply_markup=InlineKeyboardMarkup(contracts.CONTRACT_EDITING_OPTIONS))

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_CHOOSING_WHAT_IN_CONTRACT_TO_EDIT:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        contract_id = gv.CALLBACK_DATA['contract_id']
        if query.data == "отмена":
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
            await context.bot.send_message(GROUP_ID, '🆗')
        elif query.data == 'payment_schedule':
            await contracts.edit_contract_payment_schedule(update, context)
        elif query.data == 'change_document_itself':
            gv.set_group_status('accountant', gv.Status_SENDING_UPDATED_CONTRACT)
            await context.bot.send_message(GROUP_ID, 'Скинь обновленный договор ⬇', reply_markup=ReplyKeyboardMarkup(gv.CANCEL_BUTTON))
        elif query.data == 'close_contract':
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
            db.mark_contract_closed(contract_id)
            await context.bot.send_message(GROUP_ID, 'Закрыл 👍')

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_BUSINESS_DIRECTION_TO_SEND_CONTRACT:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
            await context.bot.send_message(GROUP_ID, '🆗')
        else:
            gv.set_group_status('accountant', gv.Status_PRESSING_CLIENT_TO_SEND_CONTRACT)
            buttons = []
            clients_ = db.get_clients_having_active_contracts_from_business_direction(gv.CALLBACK_DATA[query.data])
            gv.CALLBACK_DATA.clear()
            for client in clients_:
                idd = client['id']
                surname = client['surname']
                name = client['name']
                phone_number = client['phone_number']
                buttons.append([InlineKeyboardButton(f'{name} {surname}, {prettify_phone_number(phone_number)}', callback_data=f'{idd}')])
            buttons.append([InlineKeyboardButton("ни с кем, забей", callback_data="отмена")])
            await context.bot.send_message(GROUP_ID, 'Договор с кем скинуть?', reply_markup=InlineKeyboardMarkup(buttons))

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_CLIENT_TO_SEND_CONTRACT:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if query.data == "отмена":
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
            await context.bot.send_message(GROUP_ID, '🆗')
        else:
            contracts_ = db.get_client_contracts(int(query.data))
            if contracts_ is None:
                logging.error("client can't have no contracts")
                exit(0)
            gv.CALLBACK_DATA.clear()
            if len(contracts_) == 1:
                gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
                await context.bot.send_document(GROUP_ID, f'{gv.PATH_CONTRACTS}/{contracts_[0][1]}')
            else:
                gv.set_group_status('accountant', gv.Status_CHOOSING_WHICH_CONTRACT_TO_SEND)
                buttons = []
                for contract in contracts_:
                    idd, name, typee = contract
                    buttons.append([InlineKeyboardButton(name, callback_data=f'{idd}')])
                    gv.CALLBACK_DATA[f'{idd}'] = name
                buttons.append([InlineKeyboardButton("никакой, забей", callback_data="отмена")])
                await context.bot.send_message(GROUP_ID, 'Какой из договоров скинуть?', reply_markup=InlineKeyboardMarkup(buttons))

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_CHOOSING_WHICH_CONTRACT_TO_SEND:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
        if query.data == "отмена":
            await context.bot.send_message(GROUP_ID, '🆗')
        else:
            await context.bot.send_document(GROUP_ID, f'{gv.PATH_CONTRACTS}/{gv.CALLBACK_DATA[query.data]}')

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_IN_SECTION_PAYMENTS:
        if query.data == 'add_paid_payment':
            await payments.add_paid_payment(update, context)
        elif query.data == 'scheduled_payments_for_today_tomorrow':
            GROUP_ID = gv.ACCOUNTANT_GROUP_ID
            msg, remark = payments.get_scheduled_payments_for_today_tomorrow_msg()
            if msg:
                await context.bot.send_message(GROUP_ID, msg, parse_mode='markdown')
                if remark:
                    await context.bot.send_message(GROUP_ID, remark)
            else:
                await context.bot.send_message(GROUP_ID, 'По расписанию платежей на сегодня и на завтра нет')
        elif query.data == 'payment_schedule':
            await payments.payment_schedule(update, context)
        elif query.data == 'list_of_clients_debts':
            await payments.list_of_clients_debts(update, context)
        elif query.data == 'list_of_actual_payments':
            await payments.list_of_actual_payments(update, context)

    elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_ADDING_PAID_PAYMENT_CONFIRMATION:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        gv.set_group_status('accountant', gv.Status_IN_SECTION_PAYMENTS)
        if query.data == 'no':
            await context.bot.send_message(GROUP_ID, '🆗', reply_markup=ReplyKeyboardRemove())
        elif query.data == 'yes':
            contract_code = gv.CALLBACK_DATA['contract_code']
            payment_date = gv.CALLBACK_DATA['payment_date']
            payment_amount = gv.CALLBACK_DATA['payment_amount']
            db.add_paid_payment(contract_code, payment_date, payment_amount)
            await context.bot.send_message(GROUP_ID, 'Добавил ✅', reply_markup=ReplyKeyboardRemove())


async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if chat_id not in [gv.ADMIN_GROUP_ID, gv.ACCOUNTANT_GROUP_ID]:
        return
    if chat_id == gv.ACCOUNTANT_GROUP_ID and gv.ACCOUNTANT_GROUP_STATUS not in [gv.Status_MAIN, gv.Status_IN_SECTION_CLIENTS,
                                                                                gv.Status_IN_SECTION_CONTRACTS, gv.Status_IN_SECTION_PAYMENTS,
                                                                                gv.Status_SENDING_NEW_CONTRACT_TEMPLATE,
                                                                                gv.Status_ENTERING_CONTRACT_INFO_STAGE_2,
                                                                                gv.Status_SENDING_UPDATED_CONTRACT]:
        await context.bot.send_message(gv.ACCOUNTANT_GROUP_ID, 'Сначала заверши предыдущее действие, нажав одну из кнопок')
        return

    if chat_id == gv.ACCOUNTANT_GROUP_ID:
        GROUP_ID = gv.ACCOUNTANT_GROUP_ID
        if gv.ACCOUNTANT_GROUP_STATUS == gv.Status_SENDING_NEW_CONTRACT_TEMPLATE:
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
            contract_template_names = os.listdir(gv.PATH_CONCTRACT_TEMPLATES)
            file_name = update.message.document.file_name
            file_id = update.message.document.file_id
            tgFileInstance = await context.bot.get_file(file_id)
            if file_name in contract_template_names:
                os.remove(f'{gv.PATH_CONCTRACT_TEMPLATES}/{file_name}')
                await tgFileInstance.download_to_drive(f'{gv.PATH_CONCTRACT_TEMPLATES}/{file_name}')
                await context.bot.send_message(GROUP_ID, 'Заменил шаблон договора на сервере 👌', reply_markup=ReplyKeyboardRemove())
            else:
                await tgFileInstance.download_to_drive(f'{gv.PATH_CONCTRACT_TEMPLATES}/{file_name}')
                await context.bot.send_message(GROUP_ID, 'Добавил шаблон договора на сервер 👌', reply_markup=ReplyKeyboardRemove())

        elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_ENTERING_CONTRACT_INFO_STAGE_2:
            file_name = update.message.document.file_name
            if not file_name.startswith('Договор '):
                msg = 'Название файла должно быть подобно следующему: "Договор А1 от 03.04.2024г." Повтори попытку или нажми "отмена"'
                await context.bot.send_message(GROUP_ID, msg)
                return
            if contracts.contract_name_has_english_letters(file_name[:file_name.rfind('.')]):
                msg = 'Название договора не может содержать английские буквы. Повтори попытку или нажми "отмена"'
                await context.bot.send_message(GROUP_ID, msg)
                return
            existing_contract_names = os.listdir(gv.PATH_CONTRACTS)
            if file_name in existing_contract_names:
                await context.bot.send_message(GROUP_ID, 'Договор с таким названием уже существует. Выбери другое имя или нажми "отмена"')
                return
            contract_code = file_name[file_name.find(' ') + 1:]
            contract_code = contract_code[:contract_code.find(' ')].strip()
            if db.contract_code_exists(contract_code):
                msg = f'Договор с номером "{contract_code}" уже существует. Повтори попытку или нажми "отмена"'
                await context.bot.send_message(GROUP_ID, msg)
                return
            gv.CALLBACK_DATA['file_name'] = file_name
            gv.CALLBACK_DATA['contract_code'] = contract_code
            gv.CALLBACK_DATA['contract_type'] = 'regular' if file_name[8].lower() == 'а' else 'onetime'
            gv.CURRENT_FILE = f'File/{file_name}'
            file_id = update.message.document.file_id
            tgFileInstance = await context.bot.get_file(file_id)
            await tgFileInstance.download_to_drive(gv.CURRENT_FILE)
            gv.set_group_status('accountant', gv.Status_ENTERING_CONTRACT_INFO_STAGE_3)
            msg = 'Этап 3️⃣. Введи плановые даты и суммы платежей в следующем формате:\n\n'
            if gv.CALLBACK_DATA['contract_type'] == 'regular':
                msg += 'Число месяца от 1 до 28 - сумма платежа'
            else:
                msg += 'Дата в формате "дд.мм.гг." - сумма платежа'
            msg += '\n\nСколько платежей, столько должно быть и строк ввода'
            await context.bot.send_message(GROUP_ID, msg)

        elif gv.ACCOUNTANT_GROUP_STATUS == gv.Status_SENDING_UPDATED_CONTRACT:
            file_name = update.message.document.file_name
            if file_name != gv.CALLBACK_DATA['contract_file_name']:
                await context.bot.send_message(GROUP_ID, 'Имя обновленного файла не совпадает с именем старого. Повтори попытку или нажми кнопку "отмена"')
                return
            gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
            gv.CURRENT_FILE = f'File/{file_name}'
            file_id = update.message.document.file_id
            tgFileInstance = await context.bot.get_file(file_id)
            await tgFileInstance.download_to_drive(gv.CURRENT_FILE)
            os.replace(gv.CURRENT_FILE, f'{gv.PATH_CONTRACTS}/{file_name}')
            await context.bot.send_message(GROUP_ID, 'Заменил файл договора 👌', reply_markup=ReplyKeyboardRemove())

        else:  # отправка файла клиенту
            business_directions = db.get_business_directions('non-empty')
            if business_directions is None:
                await context.bot.send_message(GROUP_ID, 'Некому слать файл, в базе пока что никого нет')
                return
            file_id = update.message.document.file_id
            gv.CURRENT_FILE = f'File/{update.message.document.file_name}'
            tgFileInstance = await context.bot.get_file(file_id)
            await tgFileInstance.download_to_drive(gv.CURRENT_FILE)

            gv.set_group_status('accountant', gv.Status_PRESSING_BUSINESS_DIRECTION_TO_SEND_FILE)
            gv.CALLBACK_DATA.clear()
            buttons = []
            for i in range(len(business_directions)):
                buttons.append([InlineKeyboardButton(business_directions[i], callback_data=f"btn_{i}")])
                gv.CALLBACK_DATA[f"btn_{i}"] = business_directions[i]
            buttons.append([InlineKeyboardButton("никому, забей", callback_data="отмена")])
            await context.bot.send_message(GROUP_ID, 'Кому отправить этот файл? Выбери направление бизнеса', reply_markup=InlineKeyboardMarkup(buttons))

    elif chat_id == gv.ADMIN_GROUP_ID:
        GROUP_ID = gv.ADMIN_GROUP_ID
        if update.message.document.file_name[-5:] != '.xlsx':
            await context.bot.send_message(GROUP_ID, 'Ты прикалываешься? У файла должно быть расширение ".xlsx" 🤦‍♂️')
            return
        try:
            file_path = update.message.document.file_name
            datetime.strptime(file_path[:file_path.rfind('.')], '%d%m%y').date()
        except ValueError:
            await context.bot.send_message(GROUP_ID, 'Некорректное название файла. Должна быть существующая дата в формате "ддммгг"')
            return
        gv.CURRENT_FILE = f'File/{update.message.document.file_name}'
        file_id = update.message.document.file_id
        tgFileInstance = await context.bot.get_file(file_id)
        await tgFileInstance.download_to_drive(gv.CURRENT_FILE)

        error_log = db.write_payments_data(gv.CURRENT_FILE)
        os.remove(gv.CURRENT_FILE)
        if error_log:
            await context.bot.send_message(GROUP_ID, f'Не удалось прочитать файл (\n\n{error_log}')
            return

        informator = 'informator'
        current_jobs = context.job_queue.get_jobs_by_name(informator)
        if current_jobs:
            for job in current_jobs:
                job.schedule_removal()
        context.job_queue.run_once(inform_accountant_about_new_debts, 2, name=informator)

        await context.bot.send_message(GROUP_ID, 'Записал в базу данных ✅')


async def inform_accountant_about_new_debts(context: CallbackContext):
    msg = debt_display_handler.get_unpaid_shifts_message()
    if msg:
        msg = f'У нас появились новые задолженности. Вот все они в одном списке 👇\n\n' + msg
        if gv.ACCOUNTANT_GROUP_STATUS in [gv.Status_MAIN, gv.Status_PRESSING_BUTTON_TO_DECIDE_WHETHER_TO_PAY_SALARIES]:
            gv.set_group_status('accountant', gv.Status_PRESSING_BUTTON_TO_DECIDE_WHETHER_TO_PAY_SALARIES)
            buttons = [[InlineKeyboardButton('Не сейчас 🧑‍💻', callback_data='не сейчас'),
                        InlineKeyboardButton('Оплатить 🏃‍♂️', callback_data='оплатить')]]
            await context.bot.send_message(gv.ACCOUNTANT_GROUP_ID, msg, reply_markup=InlineKeyboardMarkup(buttons))
        else:
            await context.bot.send_message(gv.ACCOUNTANT_GROUP_ID, msg)
    else:
        await context.bot.send_message(gv.ADMIN_GROUP_ID, 'Не нашел новых задолженностей в файле, ты проверял меня, да? 😎')


async def debt_reminder(context: CallbackContext):
    msg = debt_display_handler.get_unpaid_shifts_message()
    if msg:
        msg = 'Хочу напомнить, что на данный момент у нас есть некоторые задолженности перед работниками склада, было бы неплохо их погасить) 👇\n\n' + msg
        await context.bot.send_message(gv.ACCOUNTANT_GROUP_ID, msg)


async def today_tomorrow_scheduled_payments_reminder(context: CallbackContext):
    msg, remark = payments.get_scheduled_payments_for_today_tomorrow_msg()
    if msg:
        await context.bot.send_message(gv.ACCOUNTANT_GROUP_ID, msg, parse_mode='markdown')
        if remark:
            await context.bot.send_message(gv.ACCOUNTANT_GROUP_ID, remark)


async def command_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.chat.type == 'private':  # если написал простой смертный
        tg_id = update.effective_chat.id
        if db.client_exists(tg_id):
            await update.message.reply_text('Приветствую! Твоя подписка подтверждена)')
            client_info = db.get_clients_data(update.effective_user.id)
            if client_info:
                surname, name, patronymic, sex, tg_id, inn, bd_name, phone_number, email = client_info
                msg = f'Хорошие новости! Наш клиент, {name} {surname} [{prettify_phone_number(phone_number)}], только что написал мне и теперь мы можем отправлять ему файлы 😃'
                await context.bot.send_message(gv.ACCOUNTANT_GROUP_ID, msg)
        else:
            await update.message.reply_text('Приветствую! Если ты оформил подписку, то тебя пока не добавили в базу. Как только добавят, я сразу сообщу ;)')
    elif update.effective_chat.id == gv.ACCOUNTANT_GROUP_ID:
        if not context.job_queue.jobs():
            context.job_queue.run_daily(debt_reminder, time(17, 0, 0, 0))
            context.job_queue.run_daily(today_tomorrow_scheduled_payments_reminder, time(12, 0, 0, 0))
            await context.bot.send_message(gv.ACCOUNTANT_GROUP_ID, '👌')
        else:
            await context.bot.send_message(gv.ACCOUNTANT_GROUP_ID, 'Ты мне не нравишься')


async def command_pay_salaries(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if gv.ACCOUNTANT_GROUP_STATUS == gv.Status_PRESSING_BUTTON_TO_DECIDE_WHETHER_TO_PAY_SALARIES:
        gv.set_group_status('accountant', gv.Status_MAIN)
    GROUP_ID = gv.ACCOUNTANT_GROUP_ID
    if update.effective_chat.id != GROUP_ID:
        return
    if gv.ACCOUNTANT_GROUP_STATUS not in [gv.Status_MAIN, gv.Status_IN_SECTION_CLIENTS, gv.Status_IN_SECTION_CONTRACTS, gv.Status_IN_SECTION_PAYMENTS]:
        await context.bot.send_message(GROUP_ID, 'Сначала заверши предыдущее действие, нажав одну из кнопок')
        return

    msg = debt_display_handler.get_unpaid_shifts_message()
    if msg:
        msg = f'Отличная идея! Вот список задолженностей на данный момент 👇\n\n' + msg
        await context.bot.send_message(gv.ACCOUNTANT_GROUP_ID, msg)
    else:
        await context.bot.send_message(GROUP_ID, 'Никаких задолженностей на текущий момент нет!')
        return

    gv.CURRENT_OFFSET = 0
    gv.set_group_status('accountant', gv.Status_PRESSING_BUTTON_WHILE_PAYING_SALARIES)
    msg, gv.CURRENT_DEBT = debt_display_handler.get_next_payment_message()
    await context.bot.send_message(GROUP_ID, 'Начнем с начала ⬇️')
    await context.bot.send_message(GROUP_ID, msg, parse_mode='markdown', reply_markup=InlineKeyboardMarkup(gv.PAY_SALARIES_INLINE_KEYBOARD))


async def command_delete_sent_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    GROUP_ID = gv.ACCOUNTANT_GROUP_ID
    if update.effective_chat.id != GROUP_ID:
        return
    if gv.ACCOUNTANT_GROUP_STATUS not in [gv.Status_MAIN, gv.Status_IN_SECTION_CLIENTS, gv.Status_IN_SECTION_CONTRACTS, gv.Status_IN_SECTION_PAYMENTS]:
        await context.bot.send_message(GROUP_ID, 'Сначала заверши предыдущее действие, нажав одну из кнопок')
        return

    last_sent_messages = db.get_last_sent_messages()
    if last_sent_messages is None:
        await context.bot.send_message(GROUP_ID, 'За последние 2 суток мы не отправляли сообщений клиентам. Сообщения, отправленные еще раньше уже невозможно удалить')
        return
    gv.set_group_status('accountant', gv.Status_PRESSING_CLIENT_TO_DELETE_LAST_MESSAGE)
    gv.CALLBACK_DATA.clear()
    buttons = []
    for i in range(len(last_sent_messages)):
        tg_id, message_id, name, surname, phone_number = last_sent_messages[i]
        buttons.append([InlineKeyboardButton(f'{name} {surname}, {prettify_phone_number(phone_number)}', callback_data=f'{tg_id},{message_id}')])
        gv.CALLBACK_DATA[f'{tg_id},{message_id}'] = f'{name} {surname}'
    buttons.append([InlineKeyboardButton("Ни у кого не надо", callback_data="отмена")])
    await context.bot.send_message(GROUP_ID, 'У кого из них удалить последнее сообщение?', reply_markup=InlineKeyboardMarkup(buttons))


async def command_check_last_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    GROUP_ID = gv.ACCOUNTANT_GROUP_ID
    if update.effective_chat.id != GROUP_ID:
        return

    last_message = db.get_last_sent_messages(latest_one=True)
    if last_message is None:
        await context.bot.send_message(GROUP_ID, 'Мы либо пока не отправляли ничего клиентам, либо все сообщения удалили)')
        return
    tg_id, message_id, name, surname, phone_number, sending_datetime = last_message
    await context.bot.send_message(GROUP_ID, f'Последнее сообщение мы отправили клиенту по имени {name} {surname}, {sending_datetime.strftime("%d.%m.%y в %H:%M")}. Пересылаю его сюда ⬇️')
    try:
        await context.bot.forward_message(GROUP_ID, tg_id, message_id)
    except Exception as e:
        logging.error(str(e))
        await context.bot.send_message(GROUP_ID, f'Что-то пошло не так...\n{str(e)}')


async def command_get_shifts_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    GROUP_ID = gv.ACCOUNTANT_GROUP_ID
    if update.effective_chat.id != GROUP_ID:
        return

    shifts_data = db.get_all_shifts()
    if shifts_data is None:
        await context.bot.send_message(GROUP_ID, 'У нас пока не было задолженностей)')
        return
    prettify_shifts_data(shifts_data)
    df = pd.DataFrame(shifts_data, columns=['Дата', 'Имя', 'Телефон', 'Баркод', 'Название', 'Тариф', 'Количество', 'Акт', 'Оплачено'])
    file_path = f'File/shifts_data.xlsx'
    df.to_excel(file_path, sheet_name='Смены', index=False)
    prettify_excel_columns(file_path)
    await context.bot.send_document(GROUP_ID, file_path)
    os.remove(file_path)


async def command_clients(update: Update, context: ContextTypes.DEFAULT_TYPE):
    GROUP_ID = gv.ACCOUNTANT_GROUP_ID
    if update.effective_chat.id != GROUP_ID:
        return
    if gv.ACCOUNTANT_GROUP_STATUS not in [gv.Status_MAIN, gv.Status_IN_SECTION_CLIENTS, gv.Status_IN_SECTION_CONTRACTS, gv.Status_IN_SECTION_PAYMENTS]:
        await context.bot.send_message(GROUP_ID, 'Сначала заверши предыдущее действие, нажав одну из кнопок')
        return

    gv.set_group_status('accountant', gv.Status_IN_SECTION_CLIENTS)
    buttons = [[InlineKeyboardButton('Добавить направление бизнеса', callback_data='add_business_direction')],
               [InlineKeyboardButton('Добавить клиента', callback_data='add_client')],
               [InlineKeyboardButton('Получить клиентские данные', callback_data='get_clients_data')],
               [InlineKeyboardButton('Редактировать данные клиента', callback_data='edit_client')],
               [InlineKeyboardButton('Удалить клиента', callback_data='delete_client')],
               [InlineKeyboardButton('Удалить направление бизнеса', callback_data='delete_business_direction')],
               [InlineKeyboardButton('Показать текущие направления бизнеса', callback_data='show_business_directions')]]
    await context.bot.send_message(GROUP_ID, 'Что сделать?', reply_markup=InlineKeyboardMarkup(buttons))


async def command_contracts(update: Update, context: ContextTypes.DEFAULT_TYPE):
    GROUP_ID = gv.ACCOUNTANT_GROUP_ID
    if update.effective_chat.id != GROUP_ID:
        return
    if gv.ACCOUNTANT_GROUP_STATUS not in [gv.Status_MAIN, gv.Status_IN_SECTION_CLIENTS, gv.Status_IN_SECTION_CONTRACTS,
                                          gv.Status_IN_SECTION_PAYMENTS]:
        await context.bot.send_message(GROUP_ID, 'Сначала заверши предыдущее действие, нажав одну из кнопок')
        return

    gv.set_group_status('accountant', gv.Status_IN_SECTION_CONTRACTS)
    buttons = [[InlineKeyboardButton('Скинуть шаблон договора', callback_data='send_contract_template')],
               [InlineKeyboardButton('Добавить договор', callback_data='add_contract')],
               [InlineKeyboardButton('Список договоров', callback_data='list_of_contracts')],
               [InlineKeyboardButton('Скинуть договор', callback_data='send_contract')],
               [InlineKeyboardButton('Редактировать договор', callback_data='edit_contract')],
               [InlineKeyboardButton('Добавить/заменить шаблон договора', callback_data='add_or_change_contract_template')],
               [InlineKeyboardButton('Удалить шаблон договора', callback_data='delete_contract_template')]]
    await context.bot.send_message(GROUP_ID, 'Что сделать?', reply_markup=InlineKeyboardMarkup(buttons))


async def command_payments(update: Update, context: ContextTypes.DEFAULT_TYPE):
    GROUP_ID = gv.ACCOUNTANT_GROUP_ID
    if update.effective_chat.id != GROUP_ID:
        return
    if gv.ACCOUNTANT_GROUP_STATUS not in [gv.Status_MAIN, gv.Status_IN_SECTION_CLIENTS, gv.Status_IN_SECTION_CONTRACTS,
                                          gv.Status_IN_SECTION_PAYMENTS]:
        await context.bot.send_message(GROUP_ID, 'Сначала заверши предыдущее действие, нажав одну из кнопок')
        return

    gv.set_group_status('accountant', gv.Status_IN_SECTION_PAYMENTS)
    buttons = [[InlineKeyboardButton('Добавить оплаченный платеж', callback_data='add_paid_payment')],
               [InlineKeyboardButton('Плановые платежи на сегодня-завтра', callback_data='scheduled_payments_for_today_tomorrow')],
               [InlineKeyboardButton('График платежей', callback_data='payment_schedule')],
               [InlineKeyboardButton('Список задолженностей клиентов', callback_data='list_of_clients_debts')],
               [InlineKeyboardButton('Список фактических платежей', callback_data='list_of_actual_payments')]]
    await context.bot.send_message(GROUP_ID, 'Что сделать?', reply_markup=InlineKeyboardMarkup(buttons))


def prettify_phone_number(phone_number: str) -> str:
    assert len(phone_number) >= 9
    phone_number = phone_number[:1] + ' (' + phone_number[1:4] + ') ' + phone_number[4:7] + '-' + phone_number[7:9] + '-' + phone_number[9:]
    return phone_number


def prettify_shifts_data(shifts_data: list):
    for line in shifts_data:
        line[0] = line[0].strftime("%d.%m.%Y")
        line[2] = prettify_phone_number(line[2])
        line[-1] = 'да' if line[-1] is True else 'нет'


def prettify_excel_columns(file_path: str):
    wb = xl.load_workbook(file_path)
    sheet = wb.active
    money_format = '_-* #,##0\\ "₽"_-;\\-* #,##0\\ "₽"_-;_-* "-"\\ "₽"_-;_-@_-'

    for i in range(2, sheet.max_row + 1):
        sheet.cell(i, 6).number_format = money_format  # Тариф

    for i in range(1, sheet.max_column + 1):
        sheet.cell(1, i).fill = GradientFill(stop=("E2EFDA", "E2EFDA"))

    wb.save(file_path)


async def error(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logging.error(context.error)


def extract_status_change(chat_member_update: ChatMemberUpdated) -> Optional[Tuple[bool, bool]]:
    """Takes a ChatMemberUpdated instance and extracts whether the 'old_chat_member' was a member
    of the chat and whether the 'new_chat_member' is a member of the chat. Returns None, if
    the status didn't change.
    """
    status_change = chat_member_update.difference().get("status")
    old_is_member, new_is_member = chat_member_update.difference().get("is_member", (None, None))

    if status_change is None:
        return None

    old_status, new_status = status_change
    was_member = old_status in [
        ChatMember.MEMBER,
        ChatMember.OWNER,
        ChatMember.ADMINISTRATOR,
    ] or (old_status == ChatMember.RESTRICTED and old_is_member is True)
    is_member = new_status in [
        ChatMember.MEMBER,
        ChatMember.OWNER,
        ChatMember.ADMINISTRATOR,
    ] or (new_status == ChatMember.RESTRICTED and new_is_member is True)

    return was_member, is_member


async def track_chats(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Tracks the chats the bot is in."""
    result = extract_status_change(update.my_chat_member)
    if result is None:
        return
    was_member, is_member = result

    cause = update.effective_user
    chat = update.effective_chat
    if chat.type in [Chat.GROUP, Chat.SUPERGROUP]:
        if not was_member and is_member:
            logging.info(f"{cause.full_name} ({cause.id}) added the bot to the group \"{chat.title}\"")
            if 'админ' in chat.title.lower():
                ADMIN_GROUP_ID = chat.id
                with open("admin_group_id.txt", "w") as f:
                    f.write(str(ADMIN_GROUP_ID))
                await context.bot.set_my_commands(commands=[], scope=BotCommandScopeChat(chat.id))
            elif 'бухгалтер' in chat.title.lower():
                gv.set_group_status('accountant', gv.Status_MAIN)
                ACCOUNTANT_GROUP_ID = chat.id
                with open("accountant_group_id.txt", "w") as f:
                    f.write(str(ACCOUNTANT_GROUP_ID))
                await context.bot.set_my_commands(commands=bot_commands_accountant, scope=BotCommandScopeChat(chat.id))
            else:
                logging.warning('В названии группы, в которую добавляется бот, должно быть либо слово "бухгалтер", либо слово "админ"')
            await context.bot.set_my_commands(commands=bot_commands_client, scope=BotCommandScopeAllPrivateChats())
        elif was_member and not is_member:
            logging.info(f"{cause.full_name} ({cause.id}) removed the bot from the group \"{chat.title}\"")


def run_bot():
    print('Starting bot...')
    defaults = Defaults(tzinfo=pytz.timezone('Europe/Moscow'))
    app = Application.builder().token(TOKEN).defaults(defaults).build()

    # Commands
    app.add_handler(CommandHandler('start', command_start))
    app.add_handler(CommandHandler('pay_salaries', command_pay_salaries))
    app.add_handler(CommandHandler('delete_sent_message', command_delete_sent_message))
    app.add_handler(CommandHandler('check_last_message', command_check_last_message))
    app.add_handler(CommandHandler('get_shifts_data', command_get_shifts_data))
    app.add_handler(CommandHandler('clients', command_clients))
    app.add_handler(CommandHandler('contracts', command_contracts))
    app.add_handler(CommandHandler('payments', command_payments))

    # Messages
    app.add_handler(MessageHandler(filters.TEXT, handle_message))

    # Files
    app.add_handler(MessageHandler(filters.ATTACHMENT, handle_file))

    # Callback buttons
    app.add_handler(CallbackQueryHandler(handle_callback))

    # Adding to group
    app.add_handler(ChatMemberHandler(track_chats, ChatMemberHandler.MY_CHAT_MEMBER))

    # Errors
    app.add_error_handler(error)

    # Pools the bot
    print('Polling...')
    app.run_polling(poll_interval=1)


if __name__ == '__main__':
    # Настройка логов
    logging.basicConfig(
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        filename='info.log',
        filemode='w',
        level=logging.INFO
    )
    logging.getLogger("httpx").setLevel(logging.WARNING)

    # Создание папки File для хранения отсылаемых файлов
    if not os.path.exists("File"):
        os.mkdir("File")
    else:
        for fname in os.listdir("File"):
            os.remove(f"File/{fname}")

    if not os.path.exists("Contracts"):
        os.mkdir("Contracts")
    if not os.path.exists("Contract_templates"):
        os.mkdir("Contract_templates")

    # Инициализация нужной стратегии для формата сообщений при оплате задолженностей
    debt_display_handler = DebtDisplayHandler(AggregationByPeople())

    # Инициализация объекта базы данных
    db = Database(db_dbname=db_dbname,
                  db_host=db_host,
                  db_user=db_user,
                  db_password=db_password)

    # Глобальные переменные
    gv = GlobalVars()

    clients = Clients(gv, db)
    contracts = Contracts(gv, db)
    payments = Payments(gv, db)

    bot_commands_accountant = [
        ('pay_salaries', 'Скинуть зарплаты'),
        ('delete_sent_message', 'Удалить отправленное сообщение'),
        ('check_last_message', 'Посмотреть последнее отправленное сообщение'),
        ('get_shifts_data', 'Получить данные всех смен'),
        ('clients', 'Пидарасы с деньгами'),
        ('contracts', 'Договоры'),
        ('payments', 'Платежи'),
    ]
    bot_commands_client = [
        ('start', 'Узнать свой статус')
    ]
    run_bot()
